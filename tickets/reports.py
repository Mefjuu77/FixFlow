"""
Moduł raportów – eksport zgłoszeń do CSV/XLSX.
Dostępny wyłącznie dla administratorów.
"""

import csv
import io
from datetime import datetime

from django.http import HttpResponse
from django.db.models import Count, Sum, Q
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response

from .models import Ticket


class ReportExportView(APIView):
    """
    GET /api/reports/export/
    Generuje raport zgłoszeń w formacie CSV lub XLSX.
    Dostępne wyłącznie dla administratorów.
    """
    permission_classes = [IsAuthenticated]

    COLUMNS = [
        ('ID', lambda t: t.id),
        ('Tytuł', lambda t: t.title),
        ('Status', lambda t: t.get_status_display()),
        ('Priorytet', lambda t: t.get_priority_display()),
        ('Kategoria', lambda t: t.category.name if t.category else ''),
        ('Zgłaszający', lambda t: f'{t.creator.first_name} {t.creator.last_name}'.strip() if t.creator else ''),
        ('E-mail zgłaszającego', lambda t: t.creator.email if t.creator else ''),
        ('Technik', lambda t: f'{t.technician.first_name} {t.technician.last_name}'.strip() if t.technician else 'Brak'),
        ('Utworzono', lambda t: t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else ''),
        ('Zaktualizowano', lambda t: t.updated_at.strftime('%Y-%m-%d %H:%M') if t.updated_at else ''),
        ('Liczba komentarzy', lambda t: getattr(t, 'comments_count', 0)),
        ('Czas pracy (min)', lambda t: getattr(t, 'work_minutes', 0) or 0),
    ]

    def _check_admin(self, request):
        if request.user.role != 'ADMIN':
            raise PermissionDenied('Tylko administrator ma dostęp do raportów.')

    def _build_queryset(self, params):
        qs = Ticket.objects.select_related('category', 'creator', 'technician')

        # Zakres dat
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        if date_from:
            try:
                qs = qs.filter(created_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
            except ValueError:
                pass
        if date_to:
            try:
                qs = qs.filter(created_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
            except ValueError:
                pass

        # Multi-value filtry
        statuses = params.getlist('status')
        if statuses:
            qs = qs.filter(status__in=statuses)

        priorities = params.getlist('priority')
        if priorities:
            qs = qs.filter(priority__in=priorities)

        categories = params.getlist('category')
        if categories:
            qs = qs.filter(category_id__in=categories)

        technicians = params.getlist('technician')
        if technicians:
            qs = qs.filter(technician_id__in=technicians)

        creators = params.getlist('creator')
        if creators:
            qs = qs.filter(creator_id__in=creators)

        # Adnotacje
        qs = qs.annotate(
            comments_count=Count('comments'),
            work_minutes=Sum('work_logs__duration_minutes'),
        )

        return qs.order_by('-created_at')

    def get(self, request):
        self._check_admin(request)

        export_format = request.query_params.get('file_format', 'csv').lower()
        qs = self._build_queryset(request.query_params)

        # Tryb podglądu – zwraca JSON z liczbą wyników i pierwszymi 10
        if request.query_params.get('preview') == '1':
            total = qs.count()
            preview_tickets = qs[:10]
            rows = []
            for t in preview_tickets:
                creator_avatar = None
                if t.creator and hasattr(t.creator, 'avatar') and t.creator.avatar:
                    creator_avatar = request.build_absolute_uri(t.creator.avatar.url)
                tech_avatar = None
                if t.technician and hasattr(t.technician, 'avatar') and t.technician.avatar:
                    tech_avatar = request.build_absolute_uri(t.technician.avatar.url)
                rows.append({
                    'id': t.id,
                    'title': t.title,
                    'status': t.get_status_display(),
                    'priority': t.get_priority_display(),
                    'category': t.category.name if t.category else '',
                    'creator': f'{t.creator.first_name} {t.creator.last_name}'.strip() if t.creator else '',
                    'creator_avatar': creator_avatar,
                    'technician': f'{t.technician.first_name} {t.technician.last_name}'.strip() if t.technician else 'Brak',
                    'technician_avatar': tech_avatar,
                    'created_at': t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
                    'comments_count': getattr(t, 'comments_count', 0),
                    'work_minutes': getattr(t, 'work_minutes', 0) or 0,
                })
            return Response({'total': total, 'preview': rows})

        # ---- Eksport CSV ----
        if export_format == 'csv':
            return self._export_csv(qs)

        # ---- Eksport XLSX ----
        if export_format == 'xlsx':
            return self._export_xlsx(qs)

        return Response({'error': 'Nieobsługiwany format. Użyj csv lub xlsx.'}, status=400)

    def _export_csv(self, qs):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="fixflow_raport_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
        # BOM for Excel compatibility
        response.write('\ufeff')

        writer = csv.writer(response, delimiter=';')
        writer.writerow([col[0] for col in self.COLUMNS])

        for ticket in qs.iterator():
            writer.writerow([col[1](ticket) for col in self.COLUMNS])

        return response

    def _export_xlsx(self, qs):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = 'Raport zgłoszeń'

        # Style nagłówków
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )

        # Nagłówki
        for col_idx, (name, _) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Dane
        even_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        data_alignment = Alignment(vertical='center', wrap_text=False)

        row_idx = 2
        for ticket in qs.iterator():
            for col_idx, (_, extractor) in enumerate(self.COLUMNS, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=extractor(ticket))
                cell.alignment = data_alignment
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = even_fill
            row_idx += 1

        # Autofit kolumn
        for col_idx in range(1, len(self.COLUMNS) + 1):
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 4, 50)

        # Zamrożenie pierwszego wiersza
        ws.freeze_panes = 'A2'

        # Autofiltr
        ws.auto_filter.ref = ws.dimensions

        # Zapis do bufora
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="fixflow_raport_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        return response
